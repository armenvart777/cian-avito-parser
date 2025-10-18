"""MAX бот — команды, отчёты, статистика, inline-кнопки (ЦИАН + Авито)."""

import json
import time
import logging
import os
from pathlib import Path

import requests
from openpyxl import load_workbook

from config import (
    MAX_BOT_TOKEN, MAX_CHAT_ID,
    EXCEL_FILE, AVITO_EXCEL_FILE,
    SEEN_IDS_FILE, AVITO_SEEN_IDS_FILE,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

API = "https://platform-api.max.ru"
HEADERS = {"Authorization": MAX_BOT_TOKEN}
USERS_FILE = "max_users.json"
OWNER_CHAT_ID = 217496156
# Список разрешённых chat_id (владелец + клиент Asdets)
ALLOWED_CHAT_IDS = {217496156, 219504951}

# === Клавиатуры ===

MAIN_KEYBOARD = {
    "type": "inline_keyboard",
    "payload": {
        "buttons": [
            [
                {"type": "callback", "text": "🟦 ЦИАН", "payload": "menu_cian"},
                {"type": "callback", "text": "🟩 Авито", "payload": "menu_avito"},
            ],
            [
                {"type": "callback", "text": "📊 Общая статистика", "payload": "stats_all"},
            ],
            [
                {"type": "callback", "text": "❓ Помощь", "payload": "help"},
            ],
        ]
    }
}

CIAN_KEYBOARD = {
    "type": "inline_keyboard",
    "payload": {
        "buttons": [
            [
                {"type": "callback", "text": "📊 Статистика", "payload": "stats_cian"},
                {"type": "callback", "text": "📋 Последние 5", "payload": "last5_cian"},
            ],
            [
                {"type": "callback", "text": "📋 Последние 10", "payload": "last10_cian"},
                {"type": "callback", "text": "📁 Excel отчёт", "payload": "report_cian"},
            ],
            [
                {"type": "callback", "text": "⬅️ Назад", "payload": "menu_main"},
            ],
        ]
    }
}

AVITO_KEYBOARD = {
    "type": "inline_keyboard",
    "payload": {
        "buttons": [
            [
                {"type": "callback", "text": "📊 Статистика", "payload": "stats_avito"},
                {"type": "callback", "text": "📋 Последние 5", "payload": "last5_avito"},
            ],
            [
                {"type": "callback", "text": "📋 Последние 10", "payload": "last10_avito"},
                {"type": "callback", "text": "📁 Excel отчёт", "payload": "report_avito"},
            ],
            [
                {"type": "callback", "text": "⬅️ Назад", "payload": "menu_main"},
            ],
        ]
    }
}


# === Утилиты ===

def load_users() -> list[dict]:
    try:
        with open(USERS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def save_users(users: list[dict]):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)


def add_user(chat_id: int, user_info: dict) -> bool:
    users = load_users()
    for u in users:
        if u["chat_id"] == chat_id:
            u["last_message"] = time.strftime("%Y-%m-%d %H:%M:%S")
            u["message_count"] = u.get("message_count", 0) + 1
            save_users(users)
            return False

    users.append({
        "chat_id": chat_id,
        "user_id": user_info.get("user_id"),
        "name": user_info.get("name", ""),
        "username": user_info.get("username", ""),
        "first_seen": time.strftime("%Y-%m-%d %H:%M:%S"),
        "last_message": time.strftime("%Y-%m-%d %H:%M:%S"),
        "message_count": 1,
    })
    save_users(users)
    return True


def send_message(chat_id: int, text: str, keyboard: dict | None = None):
    payload = {"text": text}
    if keyboard:
        payload["attachments"] = [keyboard]
    resp = requests.post(
        f"{API}/messages?chat_id={chat_id}",
        headers={**HEADERS, "Content-Type": "application/json"},
        json=payload,
        timeout=10,
    )
    if resp.status_code != 200:
        logger.error("Ошибка отправки: %d %s", resp.status_code, resp.text[:200])


def send_callback_answer(callback_id: str):
    requests.post(
        f"{API}/answers/callback?callback_id={callback_id}",
        headers={**HEADERS, "Content-Type": "application/json"},
        json={},
        timeout=5,
    )


def send_file(chat_id: int, file_path: str, caption: str = ""):
    try:
        resp = requests.post(
            f"{API}/uploads?type=file",
            headers=HEADERS,
            timeout=10,
        )
        if resp.status_code != 200:
            logger.error("Ошибка upload URL: %d %s", resp.status_code, resp.text[:200])
            return
        upload_url = resp.json().get("url")
        if not upload_url:
            logger.error("Нет URL в ответе uploads: %s", resp.text[:200])
            return

        with open(file_path, "rb") as f:
            upload_resp = requests.post(
                upload_url,
                files={"data": (os.path.basename(file_path), f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=30,
            )
        if upload_resp.status_code != 200:
            logger.error("Ошибка загрузки файла: %d %s", upload_resp.status_code, upload_resp.text[:200])
            return
        file_token = upload_resp.json().get("token")
        if not file_token:
            logger.error("Нет token: %s", upload_resp.text[:200])
            return

        payload = {
            "text": caption,
            "attachments": [{"type": "file", "payload": {"token": file_token}}],
        }
        for attempt in range(5):
            if attempt > 0:
                time.sleep(2)
            msg_resp = requests.post(
                f"{API}/messages?chat_id={chat_id}",
                headers={**HEADERS, "Content-Type": "application/json"},
                json=payload,
                timeout=10,
            )
            if msg_resp.status_code == 200:
                logger.info("Файл %s отправлен", file_path)
                break
            if "not.ready" in msg_resp.text or "not.processed" in msg_resp.text:
                logger.info("Файл обрабатывается, попытка %d/5...", attempt + 1)
                continue
            logger.error("Ошибка отправки файла: %d %s", msg_resp.status_code, msg_resp.text[:200])
            break
    except Exception as e:
        logger.error("Ошибка при отправке файла: %s", e)


# === Статистика и данные ===

def _fmt_price(p: int) -> str:
    if p >= 1_000_000:
        return f"{p / 1_000_000:.1f} млн ₽"
    return f"{p:,} ₽".replace(",", " ")


def _get_excel_stats(excel_file: str, seen_file: str) -> tuple[int, int, list[int]]:
    """Получить количество seen_ids, строк в Excel и список цен."""
    seen_count = 0
    try:
        with open(seen_file, "r") as f:
            seen_ids = json.load(f)
            seen_count = len(seen_ids)
    except (FileNotFoundError, json.JSONDecodeError):
        pass

    excel_count = 0
    prices = []
    if Path(excel_file).exists():
        try:
            wb = load_workbook(excel_file, read_only=True)
            ws = wb.active
            # Определяем индекс колонки цены
            header = [cell.value for cell in ws[1]] if ws.max_row else []
            price_col = None
            for i, h in enumerate(header):
                if h and "цена" in str(h).lower():
                    price_col = i
                    break
            if price_col is None:
                # Старый формат: цена в колонке 3 (индекс 3), новый — в 4
                price_col = 4 if "Источник" in header else 3

            for row in ws.iter_rows(min_row=2, values_only=True):
                excel_count += 1
                price_str = row[price_col] if len(row) > price_col else ""
                if price_str and isinstance(price_str, str):
                    digits = "".join(c for c in price_str if c.isdigit())
                    if digits:
                        prices.append(int(digits))
            wb.close()
        except Exception:
            pass

    return seen_count, excel_count, prices


def get_stats(source: str = "all") -> str:
    """Собрать статистику."""
    lines = []

    if source in ("all", "cian"):
        seen, excel, prices = _get_excel_stats(EXCEL_FILE, SEEN_IDS_FILE)
        lines.append("🟦 ЦИАН")
        lines.append(f"  Объявлений в базе: {seen}")
        lines.append(f"  Записей в Excel: {excel}")
        if prices:
            lines.append(f"  Цены: от {_fmt_price(min(prices))} до {_fmt_price(max(prices))}")
            lines.append(f"  Средняя: {_fmt_price(sum(prices) // len(prices))}")
        lines.append("")

    if source in ("all", "avito"):
        seen, excel, prices = _get_excel_stats(AVITO_EXCEL_FILE, AVITO_SEEN_IDS_FILE)
        lines.append("🟩 Авито")
        lines.append(f"  Объявлений в базе: {seen}")
        lines.append(f"  Записей в Excel: {excel}")
        if prices:
            lines.append(f"  Цены: от {_fmt_price(min(prices))} до {_fmt_price(max(prices))}")
            lines.append(f"  Средняя: {_fmt_price(sum(prices) // len(prices))}")

    if not lines:
        return "Данных пока нет."

    return "📊 Статистика\n\n" + "\n".join(lines)


def get_last_listings(count: int = 5, source: str = "cian") -> str:
    """Последние объявления из Excel."""
    excel_file = AVITO_EXCEL_FILE if source == "avito" else EXCEL_FILE
    source_label = "Авито" if source == "avito" else "ЦИАН"
    source_emoji = "🟩" if source == "avito" else "🟦"

    if not Path(excel_file).exists():
        return f"Excel-файл {source_label} пока пуст."

    try:
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active
        header = [cell.value for cell in ws[1]] if ws.max_row else []
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception:
        return "Ошибка чтения Excel."

    if not rows:
        return f"Объявлений {source_label} пока нет."

    # Определяем смещение колонок (новый формат с "Источник" или старый)
    has_source = "Источник" in header
    offset = 1 if has_source else 0

    last = rows[-count:]
    last.reverse()

    lines = [f"{source_emoji} Последние {min(count, len(last))} — {source_label}:\n"]
    for row in last:
        obj_type = row[1 + offset] if len(row) > 1 + offset else ""
        rooms = row[2 + offset] if len(row) > 2 + offset else ""
        price = row[3 + offset] if len(row) > 3 + offset else ""
        address = row[4 + offset] if len(row) > 4 + offset else ""
        area = row[5 + offset] if len(row) > 5 + offset else ""
        url = row[8 + offset] if len(row) > 8 + offset else ""

        if obj_type == "дом":
            title = "Дом"
        elif rooms == "студия":
            title = "Студия"
        else:
            title = f"{rooms}-комн. кв."

        line = f"• {title}, {price}"
        if area:
            line += f", {area}"
        if address:
            short_addr = address.split(",")[-1].strip() if "," in str(address) else address
            line += f"\n  {short_addr}"
        if url:
            line += f"\n  {url}"
        lines.append(line)

    return "\n".join(lines)


# === Обработка команд ===

HELP_TEXT = """🤖 Бот-парсер недвижимости (ЦИАН + Авито)

🟦 ЦИАН — статистика, последние объявления, Excel
🟩 Авито — статистика, последние объявления, Excel
📊 Общая статистика — по обоим источникам

Нажмите кнопку ниже 👇"""


def handle_command(chat_id: int, cmd: str):
    cmd = cmd.strip().lower()

    # Главное меню
    if cmd in ("menu_main", "/start", "start", "привет", "hi", "помощь", "help", "команды", "/help"):
        send_message(chat_id, HELP_TEXT, MAIN_KEYBOARD)

    # Подменю ЦИАН
    elif cmd == "menu_cian":
        send_message(chat_id, "🟦 ЦИАН — выберите действие:", CIAN_KEYBOARD)

    # Подменю Авито
    elif cmd == "menu_avito":
        send_message(chat_id, "🟩 Авито — выберите действие:", AVITO_KEYBOARD)

    # Статистика
    elif cmd == "stats_all":
        send_message(chat_id, get_stats("all"), MAIN_KEYBOARD)
    elif cmd in ("stats_cian", "статистика", "стат", "stats"):
        send_message(chat_id, get_stats("cian"), CIAN_KEYBOARD)
    elif cmd == "stats_avito":
        send_message(chat_id, get_stats("avito"), AVITO_KEYBOARD)

    # Последние ЦИАН
    elif cmd in ("last5_cian", "последние"):
        send_message(chat_id, get_last_listings(5, "cian"), CIAN_KEYBOARD)
    elif cmd == "last10_cian":
        send_message(chat_id, get_last_listings(10, "cian"), CIAN_KEYBOARD)

    # Последние Авито
    elif cmd == "last5_avito":
        send_message(chat_id, get_last_listings(5, "avito"), AVITO_KEYBOARD)
    elif cmd == "last10_avito":
        send_message(chat_id, get_last_listings(10, "avito"), AVITO_KEYBOARD)

    # Отчёты
    elif cmd in ("report_cian", "отчёт", "отчет", "excel", "файл", "report"):
        if Path(EXCEL_FILE).exists():
            send_message(chat_id, "📁 Формирую отчёт ЦИАН...")
            send_file(chat_id, EXCEL_FILE, "🟦 ЦИАН — все объявления")
        else:
            send_message(chat_id, "Excel ЦИАН пока пуст.", CIAN_KEYBOARD)

    elif cmd == "report_avito":
        if Path(AVITO_EXCEL_FILE).exists():
            send_message(chat_id, "📁 Формирую отчёт Авито...")
            send_file(chat_id, AVITO_EXCEL_FILE, "🟩 Авито — все объявления")
        else:
            send_message(chat_id, "Excel Авито пока пуст.", AVITO_KEYBOARD)

    # Текстовые команды с числом
    elif cmd.startswith("последние"):
        parts = cmd.split()
        count = 5
        if len(parts) > 1:
            try:
                count = min(int(parts[1]), 20)
            except ValueError:
                pass
        send_message(chat_id, get_last_listings(count, "cian"), CIAN_KEYBOARD)

    else:
        send_message(chat_id, "Не понял команду. Выберите действие:", MAIN_KEYBOARD)


# === Polling ===

def poll():
    marker = None
    logger.info("MAX бот запущен (ЦИАН + Авито), жду сообщения...")

    while True:
        params = {"timeout": 30, "limit": 100, "types": "message_created,message_callback"}
        if marker:
            params["marker"] = marker

        try:
            resp = requests.get(
                f"{API}/updates",
                headers=HEADERS,
                params=params,
                timeout=35,
            )
            data = resp.json()
        except Exception as e:
            logger.error("Ошибка polling: %s", e)
            time.sleep(5)
            continue

        marker = data.get("marker")

        for update in data.get("updates", []):
            update_type = update.get("update_type")

            # Callback (кнопки)
            if update_type == "message_callback":
                callback = update.get("callback", {})
                callback_id = callback.get("callback_id", "")
                payload = callback.get("payload", "")

                chat_id = None
                for path in [
                    lambda: update.get("message", {}).get("recipient", {}).get("chat_id"),
                    lambda: update.get("chat_id"),
                    lambda: callback.get("user", {}).get("user_id"),
                ]:
                    val = path()
                    if val:
                        chat_id = val
                        break

                if callback_id:
                    send_callback_answer(callback_id)

                if not chat_id:
                    logger.warning("Callback без chat_id, пропускаю")
                    continue

                if int(chat_id) not in ALLOWED_CHAT_IDS:
                    logger.info("Доступ закрыт для chat_id: %s (callback)", chat_id)
                    send_message(int(chat_id), "⛔ Доступ закрыт.")
                    continue

                logger.info("Callback: %s (chat_id: %s)", payload, chat_id)
                handle_command(int(chat_id), payload)
                continue

            # Текстовое сообщение
            if update_type != "message_created":
                continue

            msg = update.get("message", {})
            sender = msg.get("sender", {})
            body = msg.get("body", {})
            text = body.get("text", "")
            chat_id = update.get("chat_id") or msg.get("recipient", {}).get("chat_id")

            if not chat_id:
                continue

            if int(chat_id) not in ALLOWED_CHAT_IDS:
                logger.info("Новый пользователь chat_id: %s, name: %s", chat_id, sender.get("name", ""))
                send_message(int(chat_id), f"👋 Ваш ID: {chat_id}\nОтправьте этот ID разработчику для получения доступа.")
                # Уведомить владельца
                send_message(OWNER_CHAT_ID, f"🔔 Новый пользователь!\nChat ID: {chat_id}\nИмя: {sender.get('name', '?')}")
                continue

            user_info = {
                "user_id": sender.get("user_id"),
                "name": sender.get("name", ""),
                "username": sender.get("username", ""),
            }

            add_user(int(chat_id), user_info)

            logger.info(
                "Сообщение от %s (chat_id: %s): %s",
                sender.get("name", "?"),
                chat_id,
                text[:50],
            )

            handle_command(int(chat_id), text)


if __name__ == "__main__":
    poll()
